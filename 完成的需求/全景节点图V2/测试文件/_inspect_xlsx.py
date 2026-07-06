import openpyxl
wb = openpyxl.load_workbook('需求文件/全景节点图V2/测试文件/附件13：【雄安青藤小镇书院项目】蓝城伟业项目内控计划节点分解表.xlsx', read_only=True, data_only=True)
ws = wb.active
print(f'Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}')
rows = list(ws.iter_rows(min_row=3, max_row=12, values_only=True))
for i, r in enumerate(rows):
    vals = [str(c)[:40] if c else '' for c in r]
    print(f'Row {i+3}: ' + ' | '.join(vals))
print('---')
print(f'Total data rows (after header): {ws.max_row - 3}')
wb.close()
