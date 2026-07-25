"""extract_table 工具 —— 表格结构化提取。

Excel: openpyxl（原生行列 + 合并单元格）
PDF: camelot（stream/lattice 双模式）
输出: {rows[][], headers[], sheet_name?, format}
"""

from __future__ import annotations
import logging
import os
import time

logger = logging.getLogger("emily.tool.table")

_TABLE_SCHEMA = {
    "type": "object",
    "properties": {
        "file_path": {"type": "string", "description": "Excel/PDF 文件路径"},
        "sheet": {"type": "string", "description": "Excel sheet 名（可选，默认第一个）"},
        "mode": {"type": "string", "enum": ["stream", "lattice"],
                 "description": "PDF 表格提取模式，stream=流式（默认），lattice=线框"},
    },
    "required": ["file_path"],
}

_TABLE_DESCRIPTION = (
    "从 Excel 或 PDF 文件中提取表格数据，返回行列数组。"
    "Excel 用 openpyxl 原生解析，PDF 用 camelot（stream/lattice 双模式）。"
)


async def handle_extract_table(params: dict) -> dict:
    """M14 handler：表格结构化提取。

    Args:
        params: {file_path, sheet?, mode?}
    Returns:
        {success, tables[{rows[][], headers[], sheet_name?, format}]}
    """
    started = time.monotonic()
    file_path = params.get("file_path", "")
    sheet_name = params.get("sheet", "")
    mode = params.get("mode", "stream")

    if not file_path or not os.path.exists(file_path):
        return {"success": False, "error": f"file not found: {file_path}"}

    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext in (".xlsx", ".xls"):
            result = _extract_excel(file_path, sheet_name)
        elif ext == ".pdf":
            result = _extract_pdf(file_path, mode)
        else:
            return {"success": False, "error": f"unsupported file type: {ext}"}

        result["file_path"] = file_path
        result["elapsed_ms"] = int((time.monotonic() - started) * 1000)
        return result
    except ImportError as e:
        return {"success": False, "error": str(e)}
    except Exception as e:
        logger.warning("extract_table failed: %s (%s)", file_path, e)
        return {"success": False, "error": str(e),
                "elapsed_ms": int((time.monotonic() - started) * 1000)}


def _extract_excel(file_path: str, sheet_name: str) -> dict:
    """openpyxl 提取 Excel 表格。"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl not installed (pip install openpyxl)")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames

    tables = []
    for sname in sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        rows = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        if rows:
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []
            tables.append({
                "sheet_name": sname,
                "headers": headers,
                "rows": data_rows,
                "format": "excel",
            })

    wb.close()
    return {"success": True, "tables": tables}


def _extract_pdf(file_path: str, mode: str) -> dict:
    """camelot 提取 PDF 表格。"""
    try:
        import camelot
    except ImportError:
        raise ImportError("camelot-py not installed (pip install camelot-py[cv]; apt install ghostscript)")

    try:
        tables_data = camelot.read_pdf(file_path, flavor=mode, pages="all")
    except Exception as e:
        raise RuntimeError(f"camelot PDF table extraction failed: {e}")

    tables = []
    for i, tbl in enumerate(tables_data):
        df = tbl.df
        headers = list(df.iloc[0]) if len(df) > 0 else []
        data_rows = df.iloc[1:].values.tolist() if len(df) > 1 else []
        tables.append({
            "table_index": i,
            "page": tbl.page,
            "headers": headers,
            "rows": data_rows,
            "format": mode,
            "accuracy": round(tbl.parsing_report.get("accuracy", 0), 2),
        })

    return {"success": True, "tables": tables}
